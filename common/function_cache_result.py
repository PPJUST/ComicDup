import os
import pickle
from typing import List

import lzytools
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

from common.class_comic import ComicInfoBase

CACHE_MATCH_RESULT_FOLDER = 'cache/'
CACHE_MATCH_RESULT_FILE_EXTENSION = '.pkl'


def check_cache_exist(cache_dirpath: str = CACHE_MATCH_RESULT_FOLDER):
    """检查缓存文件夹是否存在"""
    if not os.path.exists(cache_dirpath):
        os.makedirs(cache_dirpath)


def save_match_result(data: List[List[ComicInfoBase]]):
    """保存匹配结果"""
    check_cache_exist(CACHE_MATCH_RESULT_FOLDER)
    # 提取时间戳
    time_str = lzytools.time.get_current_time(_format='match result %Y%m%d %H_%M_%S')
    # 组合文件名
    filename = f'{time_str}{CACHE_MATCH_RESULT_FILE_EXTENSION}'
    # 组合缓存路径
    cache_filepath = os.path.normpath(os.path.join(CACHE_MATCH_RESULT_FOLDER, filename))
    # 存储匹配结果
    with open(cache_filepath, 'wb') as file:
        pickle.dump(data, file)


def read_match_result(filename) -> List[List[ComicInfoBase]]:
    """读取匹配结果"""
    # 组合缓存路径
    cache_filepath = os.path.normpath(os.path.join(CACHE_MATCH_RESULT_FOLDER, filename))
    # 读取匹配结果
    with open(cache_filepath, 'rb') as file:
        data = pickle.load(file)
        return data


def output_match_result_to_excel(filename_cache):
    """导出匹配结果为excel"""
    data: List[List[ComicInfoBase]] = read_match_result(filename_cache)

    # xlsx文件
    wb = Workbook()
    ws = wb.active
    ws.title = '相似组'

    # 表头
    headers = ['组索引', '文件路径', '文件类型', '文件名', '文件大小(MB)']
    ws.append(headers)
    rows = []
    # 遍历二维列表
    for index, comic_group in enumerate(data, start=1):
        for comic in comic_group:
            group_index = f'第{index}组'
            filepath = comic.filepath
            filetype = comic.filetype.text
            filename = comic.filename
            filesize = f'{round(comic.filesize_bytes / (1024 * 1024), 2)}MB'

            # 追加一行数据
            row = [group_index, filepath, filetype, filename, filesize]
            ws.append(row)

            # 设置超链接
            cell = ws.cell(row=ws.max_row, column=2)
            cell.hyperlink = Hyperlink(
                ref=cell.coordinate,
                target=filepath,
                location=None,
            )
            cell.font = cell.font.copy(color="0000FF", underline="single")

    # 导出Excel
    output_path = f'{filename_cache}.xlsx'
    wb.save(output_path)
    os.startfile(output_path)
